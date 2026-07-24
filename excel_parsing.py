"""
excel_parsing.py
────────────────
FastAPI router for Excel / CSV extraction.

Supports two rendering modes chosen automatically per sheet:
  • TABULAR   – plain GFM table  (classic spreadsheet data)
  • FORM      – smart renderer   (form/report documents with merged headers,
                                  key-value question rows, signature blocks, etc.)

The renderer is chosen by inspecting the ratio of unique-value-per-row counts:
  - If most non-empty rows have ≤ 4 distinct cell values → FORM
  - Otherwise → TABULAR

Mount in any FastAPI app:
    from excel_parsing import router as excel_router
    app.include_router(excel_router, prefix="/excel", tags=["Excel Pipeline"])
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import shutil
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Optional

from dotenv import load_dotenv
from fastapi import APIRouter, BackgroundTasks, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel

load_dotenv()

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
log = logging.getLogger("excel_pipeline")

# ── Config ─────────────────────────────────────────────────────────────────────
OUTPUT_DIR  = Path(os.getenv("EXCEL_OUTPUT_DIR", "./uploads/excel"))
STAGING_DIR = OUTPUT_DIR / "_staging"
MAX_ROWS_PER_SHEET = int(os.getenv("EXCEL_MAX_ROWS", "5000"))
MAX_COLS_PER_SHEET = int(os.getenv("EXCEL_MAX_COLS", "200"))
ALLOWED_EXT = {".xlsx", ".xlsm", ".xls", ".csv"}

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
STAGING_DIR.mkdir(parents=True, exist_ok=True)

# ── In-memory job store ────────────────────────────────────────────────────────
_jobs: dict[str, dict[str, Any]] = {}
_executor = ThreadPoolExecutor(max_workers=2)

# ── Pydantic schemas ───────────────────────────────────────────────────────────
class UploadResponse(BaseModel):
    job_id: str
    status: str
    message: str

class SheetSummary(BaseModel):
    name: str
    rows: int
    cols: int
    empty: bool
    renderer: str   # "tabular" | "form"

class JobStatus(BaseModel):
    job_id: str
    status: str
    document: Optional[str] = None
    sheets: Optional[list[SheetSummary]] = None
    total_tables: Optional[int] = None
    markdown_preview: Optional[str] = None
    markdown_path: Optional[str] = None
    semantic_json_path: Optional[str] = None
    error: Optional[str] = None

router = APIRouter()


# ══════════════════════════════════════════════════════════════════════════════
# Cell / grid helpers
# ══════════════════════════════════════════════════════════════════════════════
class _SafeEncoder(json.JSONEncoder):
    def default(self, obj):
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    return str(value).strip()


def _trim_grid(grid: list[list[str]]) -> list[list[str]]:
    """Remove fully-empty trailing rows and rightmost empty columns."""
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    if not grid:
        return grid
    max_col = max(
        (max((i for i, c in enumerate(row) if c), default=-1) for row in grid),
        default=-1,
    )
    return [] if max_col < 0 else [row[:max_col + 1] for row in grid]


def _used_width(grid: list[list[str]]) -> int:
    m = 0
    for row in grid:
        for i, c in enumerate(row):
            if c:
                m = max(m, i)
    return m + 1


def _unique_vals(row: list[str], used_width: int) -> list[str]:
    """
    Return deduplicated non-empty cell values in left-to-right order.

    When openpyxl expands merged cells via merge_map the same string is
    repeated across every cell in the merge range.  Deduplicating recovers
    the original distinct values and is the key step that enables form detection.
    """
    seen: list[str] = []
    seen_set: set[str] = set()
    for c in row[:used_width]:
        c = c.strip()
        if c and c not in seen_set:
            seen.append(c)
            seen_set.add(c)
    return seen


def _unique_vals_ex(row: list[str], used_width: int) -> tuple[list[str], set[str]]:
    """
    Like _unique_vals but also returns the set of values that appeared
    MORE THAN ONCE in the raw row — these are merge-cell expanded values.

    Used by the KV buffer so that _flush_kv_buf can distinguish:
      • merge-repeated value (count > 1) → safe to suppress in subsequent rows
      • genuine single-cell value (count == 1) → must NOT be suppressed even if
        it coincidentally repeats across rows (e.g. same author in every version)
    """
    counts: dict[str, int] = {}
    seen: list[str] = []
    seen_set: set[str] = set()
    for c in row[:used_width]:
        c = c.strip()
        if c:
            counts[c] = counts.get(c, 0) + 1
            if c not in seen_set:
                seen.append(c)
                seen_set.add(c)
    merged = {v for v, n in counts.items() if n > 1}
    return seen, merged


# ══════════════════════════════════════════════════════════════════════════════
# Form detection
# ══════════════════════════════════════════════════════════════════════════════
_ROW_EMPTY  = "empty"
_ROW_HEADER = "header"   # short full-width merged section banner  (≤ 80 chars)
_ROW_TEXT   = "text"     # long full-width merged paragraph / body text (> 80 chars)
_ROW_KV     = "kv"       # question + answer  (or side-by-side pairs)
_ROW_DATA   = "data"     # genuine tabular data row


def _classify_row(unique: list[str]) -> str:
    """Classify a row by its UNIQUE (deduplicated) non-empty values.

    n == 0        → empty
    n == 1        → header (short banner) or text (long paragraph)
    2 ≤ n ≤ 6    → KV  (label/value pairs, side-by-side pairs, or small grids)
    n >= 7        → data (dense tabular row)

    Using ≤ 6 as the KV threshold means consecutive KV rows get buffered and
    flushed as a GFM table, which correctly handles signature blocks, assessment
    grids, and actions tables regardless of individual cell lengths.
    """
    n = len(unique)
    if n == 0:
        return _ROW_EMPTY
    if n == 1:
        val = unique[0]
        return _ROW_TEXT if len(val) > 80 else _ROW_HEADER
    if n <= 6:
        return _ROW_KV      # always KV — let the kv_buf decide table vs inline
    return _ROW_DATA


def _is_form_sheet(grid: list[list[str]]) -> bool:
    """
    Return True when the sheet looks like a form / report rather than a
    flat data table.

    Uses a CONSERVATIVE n<=4 threshold (stricter than the renderer's n<=6)
    so that plain 5-column data tables (e.g. Version Control) are not
    incorrectly detected as form sheets.
    """
    if not grid:
        return False
    uw = _used_width(grid)
    form = 0
    total = 0
    for row in grid:
        uv = _unique_vals(row, uw)
        n = len(uv)
        if n == 0:
            continue
        total += 1
        if n == 1 or n <= 4:   # header/text/short-kv → form
            form += 1
    return total > 0 and (form / total) > 0.40


# ══════════════════════════════════════════════════════════════════════════════
# Markdown renderers
# ══════════════════════════════════════════════════════════════════════════════
def _row_md(cells: list[str]) -> str:
    escaped = [c.replace("|", "\\|").replace("\n", " ") for c in cells]
    return "| " + " | ".join(escaped) + " |"


def _tabular_to_markdown(grid: list[list[str]], sheet_name: str) -> str:
    """Classic GFM table: row 0 = header, rows 1-N = data.

    Before rendering, leading rows where every cell contains the same
    repeated value (artifact of merged-cell expansion on title rows) are
    stripped so they don't pollute the table header.
    """
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    if not grid:
        return f"*Sheet `{sheet_name}` is empty.*\n"

    # Strip leading rows that are entirely one repeated value (merged title rows)
    while len(grid) > 1:
        ne = [c for c in grid[0] if c]
        if ne and len(set(ne)) == 1:   # all non-empty cells are identical
            grid.pop(0)
        else:
            break
    # Also strip leading fully-blank rows
    while len(grid) > 1 and all(c == "" for c in grid[0]):
        grid.pop(0)

    lines = [_row_md(grid[0]),
             "| " + " | ".join(["---"] * len(grid[0])) + " |"]
    for row in grid[1:]:
        padded = row[:len(grid[0])] + [""] * max(0, len(grid[0]) - len(row))
        lines.append(_row_md(padded))
    return "\n".join(lines)


def _flush_kv_buf(kv_buf: list[tuple[list[str], set[str]]], parts: list[str]) -> None:
    """
    Render buffered KV rows.

    kv_buf entries are (unique_vals, merged_set) tuples where merged_set
    contains values that appeared more than once in the raw row (merged cells).

    * 1 row, 1 value  -> blockquote
    * 1 row, 2 values -> **Label:** Value
    * 2+ rows         -> GFM table with merge-cell column suppression
    """
    if not kv_buf:
        return

    uvs    = [entry[0] for entry in kv_buf]
    merges = [entry[1] for entry in kv_buf]

    if len(kv_buf) == 1:
        uv = uvs[0]
        if len(uv) == 1:
            parts.append(f"> {uv[0]}")
        elif len(uv) == 2:
            parts.append(f"**{uv[0]}:** {uv[1]}")
        else:
            pairs = []
            for i in range(0, len(uv) - 1, 2):
                pairs.append(f"**{uv[i]}:** {uv[i+1] if i+1 < len(uv) else ''}")
            if len(uv) % 2 == 1:
                pairs.append(f"**{uv[-1]}:**")
            parts.append("  ·  ".join(pairs))
    else:
        max_cols = max(len(r) for r in uvs)
        rows = [r + [""] * (max_cols - len(r)) for r in uvs]

        # Suppress values that repeat in the same column from one row to the next,
        # BUT ONLY if that value was a merge-cell value in the previous row
        # (i.e. appeared multiple times in the raw row before deduplication).
        # This preserves genuine repeated data values like the same author name.
        deduped = [rows[0][:]]
        for j in range(1, len(rows)):
            new_row = []
            for i, val in enumerate(rows[j]):
                prev = rows[j - 1][i] if i < len(rows[j - 1]) else ""
                is_merge_repeat = val and val == prev and val in merges[j - 1]
                new_row.append("" if is_merge_repeat else val)
            deduped.append(new_row)

        # If any cell was suppressed (row-span merge dedup fired), the first row
        # contains span labels that GFM would render as bold table headers.
        # Insert a hidden blank header so all content rows become plain data rows.
        any_suppressed = any(
            deduped[j][i] == "" and rows[j][i] != ""
            for j in range(1, len(rows))
            for i in range(max_cols)
        )
        if any_suppressed:
            lines = [_row_md([""] * max_cols),
                     "| " + " | ".join(["---"] * max_cols) + " |"]
            for r in deduped:          # all rows become data rows
                lines.append(_row_md(r))
        else:
            lines = [_row_md(deduped[0]),
                     "| " + " | ".join(["---"] * max_cols) + " |"]
            for r in deduped[1:]:
                lines.append(_row_md(r))
        parts.append("\n".join(lines))
        parts.append("")

    kv_buf.clear()


def _form_to_markdown(grid: list[list[str]], sheet_name: str,
                      row_span_vals: set[str] | None = None) -> str:
    """
    Smart renderer for form / report sheets.

    Row classification (after unique-value deduplication):
      EMPTY  → skipped
      HEADER → Markdown heading  (## short banner / ### sub-section)
      TEXT   → plain paragraph   (long body text)
      KV     → buffered; flushed as **Label:** Value (single) or GFM table (multiple)
      DATA   → buffered and flushed as a GFM table
    """
    if not grid:
        return f"*Sheet `{sheet_name}` is empty.*\n"

    uw = _used_width(grid)
    parts: list[str] = []
    data_buf: list[list[str]] = []   # buffer for consecutive DATA rows
    kv_buf:   list[tuple[list[str], set[str]]] = []  # (unique_vals, merged_set)
    last_line: str = ""              # suppress duplicate merged-cell rows
    _row_spans = row_span_vals or set()  # values from vertically-spanning merges
    kv_deferred = False              # empty row seen while kv_buf active (may be spacer)

    def _flush_data():
        if not data_buf:
            return
        parts.append(_tabular_to_markdown([r[:] for r in data_buf], sheet_name))
        parts.append("")
        data_buf.clear()

    def _flush_kv():
        _flush_kv_buf(kv_buf, parts)

    def _flush_all():
        _flush_kv()
        _flush_data()

    header_streak = 0

    for row in grid:
        uv  = _unique_vals(row, uw)
        cls = _classify_row(uv)

        if cls == _ROW_EMPTY:
            _flush_data()
            header_streak = 0
            if kv_buf:
                # Don't flush KV immediately — empty rows in Excel are often
                # just visual spacers between form-grid rows.  Set a flag and
                # let the next row decide whether to continue or terminate.
                kv_deferred = True
            else:
                _flush_kv()
            continue

        if cls == _ROW_HEADER:
            if kv_deferred:
                kv_deferred = False
                _flush_kv()         # genuine section break -> flush now
            else:
                _flush_all()
            header_streak += 1
            line = f"\n{'##' if header_streak == 1 else '###'} {uv[0]}\n"
            if line != last_line:
                parts.append(line)
                last_line = line

        elif cls == _ROW_TEXT:
            if kv_deferred:
                kv_deferred = False
                _flush_kv()
            else:
                _flush_all()
            header_streak = 0
            line = f"\n{uv[0]}\n"
            if line != last_line:
                parts.append(line)
                last_line = line

        elif cls == _ROW_KV:
            _flush_data()
            header_streak = 0
            last_line = ""
            if kv_deferred:
                kv_deferred = False
                # Same column count (>=3) = empty row was just a visual spacer.
                if kv_buf and len(uv) == len(kv_buf[-1][0]) and len(uv) >= 3:
                    kv_buf.append((uv, _row_spans))
                    continue
                else:
                    _flush_kv()   # different width -> end old block, start fresh
            kv_buf.append((uv, _row_spans))

        else:  # _ROW_DATA
            if kv_deferred:
                kv_deferred = False
                _flush_kv()
            _flush_kv()            # KV block ends; start/continue data block
            header_streak = 0
            last_line = ""
            data_buf.append([row[i] if i < len(row) else "" for i in range(uw)])

    _flush_all()
    return "\n".join(p for p in parts if p is not None)


# ══════════════════════════════════════════════════════════════════════════════
# Per-format sheet extractors
# ══════════════════════════════════════════════════════════════════════════════
def _extract_xlsx(file_path: Path) -> list[dict[str, Any]]:
    """Extract .xlsx/.xlsm sheets with merged-cell expansion."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: uv add openpyxl")

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=False)
    sheets = []

    for ws in wb.worksheets:
        # Build merge map: (row, col) → top-left value
        merge_map: dict[tuple[int, int], str] = {}
        for mr in ws.merged_cells.ranges:
            val = _cell_str(ws.cell(mr.min_row, mr.min_col).value)
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r, c)] = val

        grid: list[list[str]] = []
        max_r = min(ws.max_row or 0, MAX_ROWS_PER_SHEET)
        max_c = min(ws.max_column or 0, MAX_COLS_PER_SHEET)
        for row in ws.iter_rows(max_row=max_r, max_col=max_c):
            grid.append([
                merge_map.get((cell.row, cell.column), _cell_str(cell.value))
                for cell in row
            ])

        # Collect row_span_vals: values from merges that span multiple rows.
        # These appear once per row in the same column, so _unique_vals_ex cannot
        # detect them as repeated-within-row. We track them separately so the
        # KV table renderer can suppress them after the first row.
        row_span_vals: set[str] = set()
        for mr in ws.merged_cells.ranges:
            if mr.max_row > mr.min_row:          # vertically spanning merge
                val = _cell_str(ws.cell(mr.min_row, mr.min_col).value)
                if val:
                    row_span_vals.add(val)

        grid = _trim_grid(grid)
        sheets.append({
            "name": ws.title,
            "grid": grid,
            "meta": {
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "has_merged_cells": bool(ws.merged_cells.ranges),
                "row_span_vals": row_span_vals,
            },
        })

    wb.close()
    return sheets


def _extract_xls(file_path: Path) -> list[dict[str, Any]]:
    """Extract legacy .xls using xlrd."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("xlrd not installed. Run: uv add xlrd")

    wb = xlrd.open_workbook(str(file_path))
    sheets = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        grid = [
            [_cell_str(ws.cell_value(r, c)) for c in range(min(ws.ncols, MAX_COLS_PER_SHEET))]
            for r in range(min(ws.nrows, MAX_ROWS_PER_SHEET))
        ]
        grid = _trim_grid(grid)
        sheets.append({
            "name": name,
            "grid": grid,
            "meta": {"max_row": ws.nrows, "max_col": ws.ncols, "has_merged_cells": False},
        })
    return sheets


def _extract_csv(file_path: Path) -> list[dict[str, Any]]:
    """Parse a CSV as a single-sheet workbook."""
    with open(file_path, newline="", encoding="utf-8-sig") as fh:
        grid = [[_cell_str(c) for c in row]
                for i, row in enumerate(csv.reader(fh)) if i < MAX_ROWS_PER_SHEET]
    grid = _trim_grid(grid)
    return [{"name": file_path.stem, "grid": grid,
             "meta": {"max_row": len(grid),
                      "max_col": len(grid[0]) if grid else 0,
                      "has_merged_cells": False}}]


# ══════════════════════════════════════════════════════════════════════════════
# Core pipeline
# ══════════════════════════════════════════════════════════════════════════════
def _process_excel(file_path: Path) -> dict[str, Any]:
    suffix = file_path.suffix.lower()
    name   = file_path.stem
    out_dir = OUTPUT_DIR / name
    out_dir.mkdir(parents=True, exist_ok=True)

    log.info("Processing Excel: %s", file_path.name)

    try:
        if suffix in (".xlsx", ".xlsm"):
            sheets = _extract_xlsx(file_path)
        elif suffix == ".xls":
            sheets = _extract_xls(file_path)
        elif suffix == ".csv":
            sheets = _extract_csv(file_path)
        else:
            return {"success": False, "error": f"Unsupported format: {suffix}"}
    except RuntimeError as exc:
        return {"success": False, "error": str(exc)}
    except Exception as exc:
        log.error("Extraction failed: %s", exc)
        return {"success": False, "error": f"Extraction error: {exc}"}

    md_parts: list[str] = [f"# {name}\n"]
    sheet_summaries: list[dict] = []
    total_tables = 0

    for sheet in sheets:
        sname  = sheet["name"]
        grid   = sheet["grid"]
        rows   = len(grid)
        cols   = len(grid[0]) if grid else 0
        empty  = rows == 0

        # ── Choose renderer ──────────────────────────────────────────────────
        is_form   = (not empty) and _is_form_sheet(grid)
        renderer  = "form" if is_form else "tabular"
        log.info("  Sheet '%s' → %s renderer (%d rows, %d cols)", sname, renderer, rows, cols)

        md_parts.append(f"\n---\n\n## Sheet: {sname}\n")
        if not empty:
            if is_form:
                row_span_vals = sheet.get("meta", {}).get("row_span_vals", set())
                md_parts.append(_form_to_markdown(grid, sname, row_span_vals))
            else:
                md_parts.append(_tabular_to_markdown(grid, sname))
            total_tables += 1
        else:
            md_parts.append(f"*Sheet `{sname}` is empty.*")

        sheet_summaries.append({
            "name": sname,
            "rows": rows,
            "cols": cols,
            "empty": empty,
            "renderer": renderer,
            **sheet.get("meta", {}),
        })

    full_md = "\n".join(md_parts)
    md_path = out_dir / f"{name}.md"
    md_path.write_text(full_md, encoding="utf-8")

    semantic = {
        "document": file_path.name,
        "schema_version": "1.1",
        "total_sheets": len(sheets),
        "total_tables": total_tables,
        "sheets": [
            {
                "name": s["name"],
                "rows": len(s["grid"]),
                "cols": len(s["grid"][0]) if s["grid"] else 0,
                "empty": len(s["grid"]) == 0,
                "renderer": ("form" if _is_form_sheet(s["grid"]) else "tabular"),
                "meta": s.get("meta", {}),
                "data": s["grid"],
            }
            for s in sheets
        ],
    }
    sem_path = out_dir / f"{name}.semantic.json"
    sem_path.write_text(json.dumps(semantic, indent=2, cls=_SafeEncoder), encoding="utf-8")

    log.info("Done: %s (sheets=%d, tables=%d)", file_path.name, len(sheets), total_tables)

    return {
        "success": True,
        "document": file_path.name,
        "sheets": sheet_summaries,
        "total_tables": total_tables,
        "markdown": full_md,
        "markdown_path": str(md_path),
        "semantic_json_path": str(sem_path),
        "error": None,
    }


def _run_pipeline(job_id: str, file_path: Path) -> None:
    _jobs[job_id]["status"] = "processing"
    try:
        result = _process_excel(file_path)
        _jobs[job_id].update({
            "status": "done" if result["success"] else "error",
            "result": result,
        })
    except Exception as exc:
        log.exception("Excel pipeline crashed for job %s", job_id)
        _jobs[job_id].update({
            "status": "error",
            "result": {"success": False, "error": str(exc)},
        })
    finally:
        try:
            file_path.unlink(missing_ok=True)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# API endpoints
# ══════════════════════════════════════════════════════════════════════════════
@router.post("/upload", response_model=UploadResponse)
async def upload_excel(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    """Upload a .xlsx, .xlsm, .xls, or .csv file for extraction."""
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXT:
        raise HTTPException(
            status_code=400,
            detail=f"Only {', '.join(sorted(ALLOWED_EXT))} accepted. Got: '{suffix}'",
        )
    job_id = str(uuid.uuid4())
    tmp_path = STAGING_DIR / f"{job_id}{suffix}"
    try:
        with tmp_path.open("wb") as fh:
            shutil.copyfileobj(file.file, fh)
    finally:
        await file.close()

    _jobs[job_id] = {"status": "queued", "result": None, "filename": file.filename}
    background_tasks.add_task(_executor.submit, _run_pipeline, job_id, tmp_path)
    return UploadResponse(
        job_id=job_id,
        status="queued",
        message=f"'{file.filename}' queued. Poll /excel/status/{job_id}",
    )


@router.get("/status/{job_id}", response_model=JobStatus)
async def get_status(job_id: str):
    """Poll job status."""
    job = _jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found.")
    result = job.get("result") or {}
    md = result.get("markdown", "")
    raw_sheets = result.get("sheets", [])
    sheets = [SheetSummary(**s) for s in raw_sheets] if raw_sheets else None
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        document=result.get("document") or job.get("filename"),
        sheets=sheets,
        total_tables=result.get("total_tables"),
        markdown_preview=md[:3000] if md else None,
        markdown_path=result.get("markdown_path"),
        semantic_json_path=result.get("semantic_json_path"),
        error=result.get("error"),
    )


@router.get("/jobs")
async def list_jobs():
    return [
        {"job_id": jid, "status": j["status"],
         "filename": j.get("filename"),
         "sheets": len((j.get("result") or {}).get("sheets", []))}
        for jid, j in _jobs.items()
    ]


@router.get("/download/{job_id}/markdown")
async def download_markdown(job_id: str):
    job = _jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="Job not ready.")
    path = (job["result"] or {}).get("markdown_path")
    if not path or not Path(path).exists():
        raise HTTPException(status_code=404, detail="Markdown file not found.")
    return FileResponse(path, media_type="text/markdown", filename=Path(path).name)


@router.get("/download/{job_id}/semantic")
async def download_semantic(job_id: str):
    job = _jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="Job not ready.")
    path = (job["result"] or {}).get("semantic_json_path")
    if not path or not Path(path).exists():
        raise HTTPException(status_code=404, detail="Semantic JSON not found.")
    return FileResponse(path, media_type="application/json", filename=Path(path).name)


@router.delete("/jobs/{job_id}")
async def delete_job(job_id: str):
    if job_id not in _jobs:
        raise HTTPException(status_code=404, detail="Job not found.")
    _jobs.pop(job_id)
    return {"deleted": job_id}
