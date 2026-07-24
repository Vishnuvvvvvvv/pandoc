"""
Create test Excel that mimics VFMS11 structure:
Assessment Outcomes rows separated by empty spacer rows.
"""
from pathlib import Path
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Exec Summary"

# Row 1: Overall rating KV
ws.merge_cells("A1:G1"); ws["A1"] = "What is the overall Value for Money rating for this Proposition?"
ws["H1"] = "Needs Improvement"

# Row 2: empty spacer
# Row 3: Section header (merged)
ws.merge_cells("A3:H3"); ws["A3"] = "Guidance on Overall Assessment"

# Row 4: empty
# Row 5: Assessment Outcomes header
ws.merge_cells("A5:H5"); ws["A5"] = "Assessment Outcomes"

# Rows 6,8,10,12: data rows — separated by empty spacer rows (7,9,11)
data = [
    ("Target Market and Distribution Strategy", "Not Applicable",   "Investment Proposition",            "Fair"),
    ("Commission Structure and Adviser Charging","Inconclusive",    "Complaints",                        "Not Applicable"),
    ("Costs and Charges",                       "Needs Improvement","Service Delivery",                  "Not Applicable"),
    ("Customer Claims",                         "Not Applicable",   "Customer Communications",           "Needs Improvement"),
]
r = 6
for label1, val1, label2, val2 in data:
    ws.merge_cells(f"A{r}:B{r}"); ws[f"A{r}"] = label1
    ws.merge_cells(f"C{r}:D{r}"); ws[f"C{r}"] = val1
    ws.merge_cells(f"E{r}:F{r}"); ws[f"E{r}"] = label2
    ws.merge_cells(f"G{r}:H{r}"); ws[f"G{r}"] = val2
    r += 2   # leave a blank row between each data row

# Row after last data row (14): empty
# Row 15: Overall Summary header
ws.merge_cells(f"A{r}:H{r}"); ws[f"A{r}"] = "Overall Summary"

# Row r+2: long body text
ws.merge_cells(f"A{r+2}:H{r+2}")
ws[f"A{r+2}"] = (
    "The propositions this review covers are categorised as The Investment Account "
    "Proposition range, and are combined together under one assessment as they are "
    "unit-linked pension plans designed to provide benefits at retirement and on death."
)

out = Path("test-docs/vfm_spacer_rows.xlsx")
out.parent.mkdir(exist_ok=True)
wb.save(str(out))
print(f"Saved: {out}")
