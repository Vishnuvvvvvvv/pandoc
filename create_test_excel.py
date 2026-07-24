"""
create_test_excel.py
────────────────────
Generates a complex form-like .xlsx that mimics the Value-for-Money Assessment
framework visible in the screenshots:
  - Full-width section headers (merged, coloured)
  - Key-value question/answer rows
  - Side-by-side dual-KV rows
  - A small embedded data table (Actions)
  - A signature block with merged label + multiple key-value rows

Run:
    python create_test_excel.py
Output:
    test-docs/vfm_assessment_complex.xlsx
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BLUE  = PatternFill("solid", fgColor="1F3864")
MID_BLUE   = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE = PatternFill("solid", fgColor="BDD7EE")
CYAN       = PatternFill("solid", fgColor="00B0F0")
YELLOW     = PatternFill("solid", fgColor="FFD700")
GREEN      = PatternFill("solid", fgColor="70AD47")
RED_FILL   = PatternFill("solid", fgColor="FF4C4C")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD       = Font(bold=True)


def _hdr(ws, row: int, text: str, fill=DARK_BLUE, cols=("B", "G")):
    ws.merge_cells(f"{cols[0]}{row}:{cols[1]}{row}")
    c = ws[f"{cols[0]}{row}"]
    c.value, c.fill, c.font = text, fill, WHITE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")


def _kv(ws, row: int, label: str, value: str,
        label_cols=("B", "D"), val_cols=("E", "G"),
        label_fill=LIGHT_BLUE, val_fill=None):
    ws.merge_cells(f"{label_cols[0]}{row}:{label_cols[1]}{row}")
    lc = ws[f"{label_cols[0]}{row}"]
    lc.value, lc.fill = label, label_fill

    ws.merge_cells(f"{val_cols[0]}{row}:{val_cols[1]}{row}")
    vc = ws[f"{val_cols[0]}{row}"]
    vc.value = value
    if val_fill:
        vc.fill = val_fill


def _kv2(ws, row: int,
          lbl1: str, val1: str, lbl2: str, val2: str,
          vf1=None, vf2=None):
    """Side-by-side double KV pair."""
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"].value = lbl1
    ws[f"B{row}"].fill = LIGHT_BLUE
    ws[f"D{row}"].value = val1
    if vf1:
        ws[f"D{row}"].fill = vf1

    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"].value = lbl2
    ws[f"E{row}"].fill = LIGHT_BLUE
    ws[f"G{row}"].value = val2
    if vf2:
        ws[f"G{row}"].fill = vf2


def build() -> Path:
    wb = openpyxl.Workbook()

    # ── Sheet 1 : Executive Summary (form-like) ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Exec Summary"

    for col, w in zip("ABCDEFG", [3, 30, 12, 15, 30, 12, 15]):
        ws1.column_dimensions[col].width = w
    ws1.row_dimensions[1].height = 30

    _hdr(ws1, 1, "Customer Propositions Framework – Value for Money Assessment – Exec. Summary")

    ws1.merge_cells("B2:G2")
    ws1["B2"] = ("Use the drop-down boxes (+ and –) in the left-hand margin "
                 "to expand and hide guidance and relevant additional sections.")
    ws1["B2"].alignment = Alignment(wrap_text=True)

    _hdr(ws1, 3, "Background and Context")
    _hdr(ws1, 4, "Guidance on Ratings")
    _hdr(ws1, 5, "Executive Summary")
    _hdr(ws1, 6, "Overall Assessment", fill=MID_BLUE)

    _kv(ws1, 7,
        "What is the overall Value for Money rating for this Proposition?",
        "Needs Improvement", val_fill=YELLOW)

    _hdr(ws1, 8, "Guidance on Overall Assessment", fill=CYAN)
    _hdr(ws1, 9, "Assessment Outcomes",            fill=MID_BLUE)

    _kv2(ws1, 10, "Target Market and Distribution Strategy", "Not Applicable",
         "Investment Proposition", "Fair", vf2=GREEN)
    _kv2(ws1, 11, "Commission Structure and Adviser Charging", "Inconclusive",
         "Complaints", "Not Applicable", vf1=RED_FILL)
    _kv2(ws1, 12, "Costs and Charges", "Needs Improvement",
         "Service Delivery", "Not Applicable", vf1=YELLOW)
    _kv2(ws1, 13, "Customer Claims", "Not Applicable",
         "Customer Communications", "Needs Improvement", vf2=YELLOW)

    _hdr(ws1, 14, "Overall Summary", fill=LIGHT_BLUE)
    ws1["B14"].font = BOLD

    ws1.merge_cells("B15:G17")
    ws1["B15"] = ("The propositions this review covers are categorised as The Investment Account "
                  "Proposition range. They are unit-linked pension plans designed to provide "
                  "benefits at retirement and on death. They were sold by Provident Mutual as "
                  "Individual Plans and Group Schemes and administered by ReAssure on the Alpha platform.")
    ws1["B15"].alignment = Alignment(wrap_text=True)

    # ── Actions table ──────────────────────────────────────────────────────────
    _hdr(ws1, 19, "Actions", fill=MID_BLUE)

    for col, label in [("B", "Action"), ("E", "Owner"), ("F", "Target Date")]:
        ws1[f"{col}20"] = label
        ws1[f"{col}20"].fill = LIGHT_BLUE
        ws1[f"{col}20"].font = BOLD
    ws1.merge_cells("B20:D20")

    ws1.merge_cells("B21:D21")
    ws1["B21"] = "Action 1 – Investigate obtaining payment out details for future assessments to give a fuller view."
    ws1["E21"] = "Customer and Product"
    ws1["F21"] = "TBC"

    for r in range(22, 25):
        ws1.merge_cells(f"B{r}:D{r}")   # empty action rows

    # ── Signature block ────────────────────────────────────────────────────────
    ws1.merge_cells("B26:B29")
    ws1["B26"] = "Assessment Completed By"
    ws1["B26"].fill = LIGHT_BLUE
    ws1["B26"].font = BOLD
    ws1["B26"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1.merge_cells("E26:E29")
    ws1["E26"] = "Assessment Approved By"
    ws1["E26"].fill = LIGHT_BLUE
    ws1["E26"].font = BOLD
    ws1["E26"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sig_left  = [("Name", "Samuel Brett"), ("Title", "Customer and Product Management Executive"),
                 ("Date Started", "21/05/2024"), ("Date Completed", "29/05/2024")]
    sig_right = [("Name", "Bharat Aheer"), ("Title", "Transformation Lead"),
                 ("", ""), ("Date Approved", "29/05/2024")]

    for i, ((lk, lv), (rk, rv)) in enumerate(zip(sig_left, sig_right)):
        r = 26 + i
        ws1[f"C{r}"] = lk;  ws1[f"D{r}"] = lv
        ws1[f"F{r}"] = rk;  ws1[f"G{r}"] = rv

    # ── Row 31: question with "please select" answer ───────────────────────────
    _kv(ws1, 31,
        "Does this Assessment alter the timing of the next Proposition Review?",
        "Please Select")
    _hdr(ws1, 32, "If 'Yes' to Altering the next Review date", fill=CYAN)
    _kv(ws1, 33, "Rationale for Changing the next Proposition Review Date", "")
    _kv2(ws1, 34,
         "Original Proposition Review Date", "",
         "Amended Proposition Review Date", "")
    _kv(ws1, 35, "Next VfM Assessment Due", "Date")
    _kv(ws1, 36,
        "Date Completed Assessment sent to the Proposition Governance Team (COOProductGovernance@aviva.com)",
        "Date")

    # ── Sheet 2 : Plain tabular data (should NOT trigger form renderer) ────────
    ws2 = wb.create_sheet("Sales Data")
    headers = ["Region", "Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Total"]
    ws2.append(headers)
    for h in ws2[1]:
        h.fill = DARK_BLUE; h.font = WHITE_FONT

    import random, string
    random.seed(42)
    regions  = ["North", "South", "East", "West"]
    products = ["Alpha Fund", "Beta Bond", "Gamma Growth", "Delta Income"]
    for reg in regions:
        for prod in products:
            q = [random.randint(100_000, 900_000) for _ in range(4)]
            ws2.append([reg, prod] + q + [sum(q)])

    # ── Sheet 3 : Mixed – table inside a form ─────────────────────────────────
    ws3 = wb.create_sheet("Version Control")
    _hdr(ws3, 1, "Document Version Control", cols=("A", "E"))
    ws3.append([""])   # row 2 blank
    for col, label in [("A","Version"),("B","Date"),("C","Author"),("D","Reviewed By"),("E","Summary of Changes")]:
        ws3[f"{col}3"] = label
        ws3[f"{col}3"].fill = MID_BLUE
        ws3[f"{col}3"].font = WHITE_FONT

    rows = [
        ("1.0","01/01/2024","Samuel Brett","Bharat Aheer","Initial draft"),
        ("1.1","15/03/2024","Samuel Brett","Bharat Aheer","Updated Q1 data"),
        ("2.0","21/05/2024","Samuel Brett","Bharat Aheer","Final submission"),
    ]
    for r in rows:
        ws3.append(list(r))

    out = Path("test-docs/vfm_assessment_complex.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(str(out))
    print(f"Created: {out}")
    return out


if __name__ == "__main__":
    build()
